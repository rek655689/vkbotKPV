from settings import *
import vk_api
import connection
import requests
import re
import pandas as pd
from openpyxl import load_workbook
from requests_html import HTMLSession
from bs4 import BeautifulSoup
import time

seconds = time.time()
local_time = time.ctime(seconds)

vk, longpoll, config = connection.connect(vk_api, requests, time, local_time)
vk_token = (vk_api.VkApi(token=access_token)).get_api()
# vk_token - токен пользователя-админа для работы с админскими правами


url = 'https://catwar.su/ajax/login'
user_agent_val = 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) ' \
                 'Chrome/90.0.4430.93 Safari/537.36 '
session = HTMLSession()
session.headers.update({'Referer': url})
session.headers.update({'User-Agent': user_agent_val})
session.post(url, {**catwar})

# positions = {'heads': ('врачеватель', 'врачевательница', 'ученик врачевателя', 'ученица врачевателя', 'советник', 'советница'),
#              'elders': ('старейшина'),
#              'elects': ('избранник духов', 'избранница духов'),
#              'guards': ('страж', 'стражница'),
#              'hunters': ('охотник', 'охотница'),
#              'futures': ('будущий охотник', 'будущая охотница', 'будущий страж', 'будущая стражница'),
#              'others': ('котёнок', 'переходящий', 'переходящая')
#              }
members = ('будущий охотник', 'будущая охотница', 'будущий страж', 'будущая стражница')
guards = ('охотник', 'охотница', 'страж', 'стражница', 'котёнок', 'старейшина')

positions = {'members': members, 'guards': guards}

page_ids = {'members': 56341561, 'guards': 56438369}

# page_ids = {positions.get('heads'): [56490990], positions.get('elders'): [56591896],
#             positions.get('elects'): [56846221],
#             positions.get('guards'): [56807806, 56968786], positions.get('hunters'): [56807807, 56970221],
#             positions.get('futures'): [56490171], positions.get('others'): [56808867]}


def get_excel():
    wb = load_workbook('table.xlsx')
    writer = pd.ExcelWriter('table.xlsx', engine='openpyxl')
    writer.book = wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    return wb, writer


def get_key(d, value):
    for k, v in d.items():
        for x in v:
            if x == value:
                return k


def get_members(page_id):
    """Перенос таблицы из wiki-страницы в excel"""
    wb, writer = get_excel()

    sheet = wb['members']

    orig_page = vk_token.pages.get(**config, owner_id=-group_id, page_id=page_id, need_source=1)['source']
    start = orig_page.find('{|') + 3
    end = orig_page.find('|}')
    page = orig_page[start:end]

    c = 0
    i = 0
    row = 1
    lens = len(page)

    while i < lens - 1:
        if c < 3:
            start = page.find('\n', i)
            end = page.find('\n', start + 1)
            s = page[start:end]
            i = end

            if c == 0:
                row += 1
                vk_id = re.findall("[0-9]+", s)[0]
                vk_name = re.findall("\|[^0-9\]\]\[]+", s)[1][1:]
                sheet.cell(row=row, column=c + 1, value=vk_id)
                sheet.cell(row=row, column=c + 2, value=vk_name)

            if c == 1:
                name = re.findall("\[[^0-9\]\]\[]+", s)[0][1:-1]
                id = re.findall("[0-9]+", s)[0]
                sheet.cell(row=row, column=c + 2, value=name)
                sheet.cell(row=row, column=c + 3, value=id)

            if c == 2:
                url = (re.findall("\[.*", s)[0].replace(' ', ''))[1:-1]
                sheet.cell(row=row, column=c + 3, value=url)

            c += 1

        else:
            c = 0
            i += 3
    wb.save('table.xlsx')
    writer.close()


def check_members():
    """Проверка игроков в таблице на нахождение в клане, правильность должности/имени/имени в ВК"""
    wb, writer = get_excel()

    for key, value in positions.items():
        # получаем данные из таблицы с определённого листа
        sheet = wb[key]

        val = ' '
        r = 1
        while val is not None:
            # итерируемся по каждой строке, пока не достигнем пустой
            r += 1
            for i in range(1, 6):
                # итерируемся по 1-5 ячейке, получая данные чела
                val = sheet.cell(row=r, column=i).value
                if val is None:
                    break
                print(val)

                if i == 1:
                    vk_id = val
                if i == 2:
                    vk_name = val
                if i == 3:
                    name = val
                if i == 4:
                    id = val
                if i == 5:
                    url = val

            profile = session.get(url).content.decode("utf-8")
            try:
                # есл нет имени или должности, значит чел не в клане
                position = BeautifulSoup(profile, 'html.parser').find('i').text
                cw_name = BeautifulSoup(profile, 'html.parser').find('big').text
            except AttributeError:
                position, cw_name = None, None

            real_vk_name = vk.users.get(user_ids=vk_id, fields='first_name, last_name')[0]
            real_vk_name = real_vk_name['first_name'] + ' ' + real_vk_name['last_name']

            if real_vk_name != vk_name:
                sheet.cell(row=r, column=2, value=real_vk_name)

            if cw_name != name:
                sheet.cell(row=r, column=3, value=cw_name)

            # if (vk.groups.isMember(group_id=group_id, user_id=vk_id) == 0) or position is None or cw_name is None:
            if position is None or cw_name is None:
                # если чел не в клане - убираем из таблицы
                sheet.delete_rows(r)
                sheet.insert_rows(r)
                if vk.groups.isMember(group_id=group_id, user_id=vk_id) == 1:
                    # если чел при этом участник группы - кикаем
                    vk_token.groups.removeUser(group_id=group_id, user_id=vk_id)
                continue

            if position is not None and position not in value:
                pos = get_key(positions, position)
                sheet.delete_rows(r)
                sheet.insert_rows(r)
                sheet = wb[pos]
                sheet.insert_rows(r)
                sheet.cell(row=r, column=1, value=vk_id)
                sheet.cell(row=r, column=2, value=real_vk_name)
                sheet.cell(row=r, column=3, value=cw_name)
                sheet.cell(row=r, column=4, value=id)
                sheet.cell(row=r, column=5, value=url)
                sheet = wb[key]
        wb.save('table.xlsx')
        df = pd.read_excel(writer, key, dtype='object')
        df = df.sort_values("name")
        df.to_excel(writer, key, index=False)
    writer.save()
    writer.close()


def add_member(vk_id, vk_name, name, id, position):
    wb, writer = get_excel()
    sheet_name = get_key(positions, position)
    sheet = wb[sheet_name]
    sheet.insert_rows(2)
    sheet.cell(row=2, column=1, value=vk_id)
    sheet.cell(row=2, column=2, value=vk_name)
    sheet.cell(row=2, column=3, value=name)
    sheet.cell(row=2, column=4, value=id)
    sheet.cell(row=2, column=5, value=f'https://catwar.su/cat{id}')
    wb.save('table.xlsx')
    df = pd.read_excel(writer, sheet_name, dtype='object')
    df = df.sort_values("name")
    df.to_excel(writer, sheet_name, index=False)
    writer.save()
    writer.close()


def add_to_page():
    """Перенос из таблицы excel на wiki-страницы"""
    wb, writer = get_excel()
    for key, value in positions.items():
        sheet = wb[key]
        text = ''
        val = sheet.cell(row=1, column=6).value + '\n'
        text += val

        r = 1
        while val is not None:
            # итерируемся по каждой строке, пока не достигнем пустой
            r += 1
            for i in range(1, 6):
                # итерируемся по 1-5 ячейке, получая данные чела
                val = sheet.cell(row=r, column=i).value
                if val is None:
                    break
                print(val)

                if i == 1:
                    vk_id = val
                if i == 2:
                    vk_name = val
                if i == 3:
                    name = val
                if i == 4:
                    id = val
                if i == 5:
                    url = val
            text += f'|-\n| [[id{vk_id}|{vk_name}]]\n| [{name}|{id}]\n| [{url}]\n'

        val = '\n' + sheet.cell(row=1, column=7).value
        text += val

        wb.save('table.xlsx')

        vk_token.pages.save(text=text, page_id=page_ids.get(key), group_id=group_id)